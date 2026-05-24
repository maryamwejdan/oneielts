"""
Excel Manager Module
Handles all Excel operations for transaction tracking
Uses openpyxl for formatting and pandas for data operations
"""
import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import EXCEL_PATH, BACKUPS_DIR, EXCEL_BACKUP_PATTERN
from database.db_manager import DatabaseManager
from utils.logger import setup_logger
from utils.helpers import is_broker_group

logger = setup_logger("ExcelManager")


class ExcelManager:
    """
    Manages Excel workbook for MOJ transactions
    Features:
    - Auto-append new rows
    - Duplicate prevention
    - Auto-calculation of commissions
    - Daily backups
    - Professional formatting
    """

    COLUMNS = [
        "Application Number",
        "Client Name",
        "Lawyer Name",
        "WhatsApp Group",
        "Transaction Type",
        "Payment Link",
        "Date",
        "Broker Group (Yes/No)",
        "Gross Amount (AED)",
        "Broker Commission %",
        "Broker Commission (AED)",
        "Net Amount (AED)",
        "Status",
        "Language",
        "Message ID",
        "Created At"
    ]

    def __init__(self, db_manager: DatabaseManager, excel_path: Path = EXCEL_PATH):
        self.db = db_manager
        self.excel_path = excel_path
        self.backup_dir = BACKUPS_DIR
        self.backup_dir.mkdir(exist_ok=True)

        self._init_workbook()
        logger.info(f"ExcelManager initialized: {excel_path}")

    def _init_workbook(self):
        """Initialize Excel workbook if not exists"""
        if not self.excel_path.exists():
            df = pd.DataFrame(columns=self.COLUMNS)
            df.to_excel(self.excel_path, index=False, sheet_name="Transactions")
            self._apply_formatting()
            logger.info("Created new Excel workbook")

    def _apply_formatting(self):
        """Apply professional formatting to workbook"""
        try:
            wb = load_workbook(self.excel_path)
            ws = wb["Transactions"]

            # Header styling
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Set row height for header
            ws.row_dimensions[1].height = 30

            wb.save(self.excel_path)
            logger.debug("Excel formatting applied")

        except Exception as e:
            logger.error(f"Error applying formatting: {e}")

    def sync_from_database(self) -> int:
        """
        Sync all transactions from database to Excel

        Returns:
            Number of rows synced
        """
        try:
            transactions = self.db.get_all_transactions(limit=10000)

            if not transactions:
                logger.info("No transactions to sync")
                return 0

            # Convert to DataFrame
            rows = []
            for t in transactions:
                rows.append({
                    "Application Number": t.get("app_number", ""),
                    "Client Name": t.get("client_name", ""),
                    "Lawyer Name": t.get("lawyer_name", "") or "",
                    "WhatsApp Group": t.get("whatsapp_group", ""),
                    "Transaction Type": t.get("transaction_type", "e-notary"),
                    "Payment Link": t.get("payment_link", ""),
                    "Date": t.get("message_date", ""),
                    "Broker Group (Yes/No)": "Yes" if t.get("is_broker") else "No",
                    "Gross Amount (AED)": t.get("gross_amount", 0),
                    "Broker Commission %": t.get("commission_percent", 20),
                    "Broker Commission (AED)": t.get("commission_amount", 0),
                    "Net Amount (AED)": t.get("net_amount", 0),
                    "Status": t.get("status", "pending"),
                    "Language": t.get("language", ""),
                    "Message ID": t.get("message_id", ""),
                    "Created At": t.get("created_at", "")
                })

            df = pd.DataFrame(rows, columns=self.COLUMNS)

            # Write to Excel
            with pd.ExcelWriter(self.excel_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, index=False, sheet_name="Transactions")

                # Add summary sheet
                self._create_summary_sheet(writer, df)

            self._apply_formatting()

            logger.info(f"Synced {len(rows)} transactions to Excel")
            return len(rows)

        except Exception as e:
            logger.error(f"Error syncing to Excel: {e}")
            return 0

    def _create_summary_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame):
        """Create summary statistics sheet"""
        try:
            if df.empty:
                return

            # Calculate summaries
            total_gross = df["Gross Amount (AED)"].sum()
            total_commission = df["Broker Commission (AED)"].sum()
            total_net = df["Net Amount (AED)"].sum()
            total_count = len(df)

            # By lawyer
            lawyer_summary = df.groupby("Lawyer Name").agg({
                "Application Number": "count",
                "Gross Amount (AED)": "sum",
                "Broker Commission (AED)": "sum"
            }).rename(columns={
                "Application Number": "Count",
                "Gross Amount (AED)": "Total Gross",
                "Broker Commission (AED)": "Total Commission"
            }).reset_index()

            # By group
            group_summary = df.groupby("WhatsApp Group").agg({
                "Application Number": "count",
                "Gross Amount (AED)": "sum"
            }).rename(columns={
                "Application Number": "Count",
                "Gross Amount (AED)": "Total Gross"
            }).reset_index()

            # Create summary DataFrame
            summary_data = {
                "Metric": [
                    "Total Transactions",
                    "Total Gross Amount (AED)",
                    "Total Commission (AED)",
                    "Total Net Amount (AED)",
                    "Average per Transaction (AED)",
                    "Generated At"
                ],
                "Value": [
                    total_count,
                    round(total_gross, 2),
                    round(total_commission, 2),
                    round(total_net, 2),
                    round(total_gross / total_count, 2) if total_count > 0 else 0,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            summary_df = pd.DataFrame(summary_data)

            # Write sheets
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            lawyer_summary.to_excel(writer, index=False, sheet_name="By Lawyer")
            group_summary.to_excel(writer, index=False, sheet_name="By Group")

        except Exception as e:
            logger.error(f"Error creating summary: {e}")

    def add_transaction(self, data: Dict[str, Any]) -> bool:
        """
        Add single transaction to Excel

        Args:
            data: Transaction data dictionary

        Returns:
            True if added
        """
        try:
            # Read existing
            df = pd.read_excel(self.excel_path, sheet_name="Transactions")

            # Check duplicate
            if data.get("app_number") in df["Application Number"].values:
                logger.warning(f"Duplicate in Excel: {data['app_number']}")
                return False

            # Calculate financials
            gross = float(data.get("gross_amount", 0))
            commission_pct = float(data.get("commission_percent", 20.0))
            commission = gross * (commission_pct / 100)
            net = gross - commission

            # Create new row
            new_row = {
                "Application Number": data.get("app_number", ""),
                "Client Name": data.get("client_name", ""),
                "Lawyer Name": data.get("lawyer_name", "") or "",
                "WhatsApp Group": data.get("whatsapp_group", ""),
                "Transaction Type": data.get("transaction_type", "e-notary"),
                "Payment Link": data.get("payment_link", ""),
                "Date": data.get("message_date", ""),
                "Broker Group (Yes/No)": "Yes" if data.get("is_broker") else "No",
                "Gross Amount (AED)": gross,
                "Broker Commission %": commission_pct,
                "Broker Commission (AED)": commission,
                "Net Amount (AED)": net,
                "Status": data.get("status", "pending"),
                "Language": data.get("language", ""),
                "Message ID": data.get("message_id", ""),
                "Created At": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }

            # Append
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # Write back
            with pd.ExcelWriter(self.excel_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, index=False, sheet_name="Transactions")
                self._create_summary_sheet(writer, df)

            self._apply_formatting()

            logger.info(f"Added transaction to Excel: {data.get('app_number')}")
            return True

        except Exception as e:
            logger.error(f"Error adding to Excel: {e}")
            return False

    def create_backup(self) -> Optional[Path]:
        """
        Create daily backup of Excel file

        Returns:
            Path to backup file or None
        """
        try:
            if not self.excel_path.exists():
                return None

            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = EXCEL_BACKUP_PATTERN.format(date=date_str)
            backup_path = self.backup_dir / backup_name

            shutil.copy2(self.excel_path, backup_path)

            # Clean old backups (keep last 30 days)
            self._cleanup_old_backups(days=30)

            logger.info(f"Excel backup created: {backup_path}")
            return backup_path

        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            return None

    def _cleanup_old_backups(self, days: int = 30):
        """Remove backup files older than specified days"""
        try:
            cutoff = datetime.now() - timedelta(days=days)
            for file in self.backup_dir.glob("*.xlsx"):
                if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                    file.unlink()
                    logger.debug(f"Removed old backup: {file.name}")
        except Exception as e:
            logger.error(f"Error cleaning backups: {e}")

    def get_excel_path(self) -> Path:
        """Get current Excel file path"""
        return self.excel_path
